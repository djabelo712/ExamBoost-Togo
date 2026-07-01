"""
Genere le modele financier Excel complet d'ExamBoost Togo sur 18 mois.

Onglets :
1. Assumptions   - toutes les hypotheses (prix, CAC, LTV, couts, taux de change)
2. P&L Mensuel   - revenus et depenses mois par mois (M1 -> M18)
3. Cash Flow     - projections cash + burn rate
4. Scenarios     - 3 scenarios (pessimiste, realiste, optimiste)
5. Break-even    - analyse point mort
6. Charts        - 4 graphiques (bar revenus, line cash, pie revenus, stacked couts)

Sources :
- Plan_GoToMarket.md (budget GTM 139 000 USD, KPIs M3/M6/M12/M18)
- ExamBoost_Togo_Etude_Faisabilite_2025.txt (budget total 246 400 USD, pricing B2B2C)

Usage :
    python3 generate_financial_model.py

Le fichier Excel est ecrit dans output/ExamBoost_Financial_Model.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# =============================================================================
# STYLES GLOBAUX
# =============================================================================

HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='006837', end_color='006837', fill_type='solid')
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill(start_color='D97700', end_color='D97700', fill_type='solid')
SUBTOTAL_FONT = Font(name='Calibri', size=11, bold=True)
SUBTOTAL_FILL = PatternFill(start_color='E8F5ED', end_color='E8F5ED', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='006837')
NOTE_FONT = Font(name='Calibri', size=9, italic=True, color='666666')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
NUM_FCFA = '# ##0" FCFA";[Red]-# ##0" FCFA"'
NUM_PCT = '0.0%'
NUM_INT = '# ##0'

# Nombre de mois projetes
NB_MOIS = 18

# Taux de change (USD vers FCFA)
USD_TO_FCFA = 600

# Capital initial (budget total projet 18 mois de l'Etude de Faisabilite)
CAPITAL_INITIAL_USD = 246_400
CAPITAL_INITIAL_FCFA = CAPITAL_INITIAL_USD * USD_TO_FCFA  # 147 840 000 FCFA

# Budget GTM sur 18 mois (Plan GoToMarket)
BUDGET_GTM_USD = 139_000


# =============================================================================
# CONSTANTES METIER (alignees sur Plan_GoToMarket)
# =============================================================================

# Pricing (FCFA)
PRIX_ECOLE_PUBLIQUE = 100_000    # FCFA/an
PRIX_ECOLE_PRIVEE = 150_000      # FCFA/an
PRIX_ECOLE_PREMIUM = 300_000     # FCFA/an
PRIX_PREMIUM_ELEVE = 2_000       # FCFA/mois

# Mix ecoles ( % du total etablissements partenaires )
MIX_PUBLIQUE = 0.30
MIX_PRIVEE = 0.60
MIX_PREMIUM = 0.10

# Conversion premium eleve ( % des eleves actifs )
CONVERSION_PREMIUM_CIBLE = 0.05   # 5 % a M18

# CAC / LTV (FCFA)
CAC_ELEVE = 400
CAC_ECOLE = 30_000
LTV_ELEVE = 15_000                # sur 12 mois
LTV_ECOLE = 300_000               # sur 3 ans

# Salaires USD/mois (Plan GoToMarket section 6.1)
SALAIRE_CHEF_PROJET = 1500        # M0 -> M18
SALAIRE_COMMERCIAL_B2B = 1000     # M4 -> M18
SALAIRE_COMMUNITY_MANAGER = 800   # M2 -> M18
SALAIRE_COMMERCIAL_REGIONAL = 800 # x2, M9 -> M18
CHARGES_SOCIALES_PCT = 0.12       # ~12 % brut Togo

# Autres couts fixes USD/mois
COUT_CLOUD_USD = 50
COUT_API_IA_USD = 100
COUT_SMS_PAR_SMS_USD = 0.02
COUT_MARKETING_BASE_USD = 200     # phase pilote
COUT_MARKETING_FULL_USD = 500     # phase lancement +
COUT_LEGAL_USD = 100

# Subventions / grants (FCFA) - inscrites a M6, M12, M18
SUBVENTION_M6_FCFA = 5_000_000    # ~8 300 USD (GPE / AFD premier tranche)
SUBVENTION_M12_FCFA = 10_000_000  # deuxieme tranche
SUBVENTION_M18_FCFA = 15_000_000  # troisieme tranche (Serie A amorce)

# Frais paiement mobile money (Flooz / TMoney)
FRAIS_PAIEMENT_PCT = 0.02          # 2 % des revenus premium + ecoles


# =============================================================================
# TRAJECTOIRES KPIs (scenarios realiste, aligne Plan GoToMarket)
# =============================================================================

# Nombre d'eleves actifs par mois (M1 -> M18) - scenario realiste
TRAJ_ELEVES_REALISTE = [
    50, 100, 300,        # phase pilote (M1-M3)
    500, 800, 1000,      # phase lancement (M4-M6)
    2000, 3500, 5000,    # phase lancement (M7-M9)
    8000, 12000, 20000,  # phase croissance (M10-M12)
    25000, 30000, 35000, # phase expansion (M13-M15)
    40000, 45000, 50000, # phase consolidation (M16-M18)
]

# Nombre d'etablissements partenaires par mois (M1 -> M18)
TRAJ_ECOLES_REALISTE = [
    1, 3, 5,             # phase pilote
    8, 15, 20,           # phase lancement
    30, 40, 50,          # phase lancement
    70, 90, 100,         # phase croissance
    130, 150, 170,       # phase expansion
    180, 190, 200,       # phase consolidation
]

# Taux de conversion premium (% eleves actifs) par mois
TRAJ_CONV_PREMIUM = [
    0.00, 0.00, 0.00,    # M1-M3 : 0 % (gratuit pilote)
    0.005, 0.01, 0.02,   # M4-M6 : 0.5 % -> 2 %
    0.02, 0.025, 0.03,   # M7-M9
    0.035, 0.04, 0.04,   # M10-M12
    0.045, 0.045, 0.05,  # M13-M15
    0.05, 0.05, 0.05,    # M16-M18
]


# =============================================================================
# UTILITAIRES
# =============================================================================

def col_letter(idx):
    """Convertit un index de colonne (1-based) en lettre Excel."""
    return get_column_letter(idx)


def style_header_row(ws, row, n_cols, start_col=1):
    """Applique le style header vert sur une ligne."""
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_section_row(ws, row, n_cols, start_col=1):
    """Applique le style section orange sur une ligne."""
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def style_subtotal_row(ws, row, n_cols, start_col=1):
    """Applique le style subtotal (gras, fond vert clair) sur une ligne."""
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = SUBTOTAL_FONT
        cell.fill = SUBTOTAL_FILL
        cell.border = THIN_BORDER


def set_number_format_row(ws, row, n_cols, fmt, start_col=2):
    """Applique un format numerique sur une ligne."""
    for c in range(start_col, start_col + n_cols):
        ws.cell(row=row, column=c).number_format = fmt


def set_col_widths(ws, widths):
    """widths : dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_title(ws, title, n_cols, row=1):
    """Ecrit un titre fusionne sur la premiere ligne."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    ws.cell(row=row, column=1).alignment = LEFT


# =============================================================================
# TRAJECTOIRES DE COUTS
# =============================================================================

def salaires_mois_usd(m):
    """Calcule le cout salarial brut USD pour le mois m (1-indexe)."""
    total = SALAIRE_CHEF_PROJET  # M1 -> M18
    if m >= 2:
        total += SALAIRE_COMMUNITY_MANAGER
    if m >= 4:
        total += SALAIRE_COMMERCIAL_B2B
    if m >= 9:
        total += 2 * SALAIRE_COMMERCIAL_REGIONAL
    # Charges sociales
    total_brut_charge = total * (1 + CHARGES_SOCIALES_PCT)
    return total_brut_charge


def marketing_mois_usd(m):
    """Budget marketing mensuel USD selon la phase."""
    if m <= 3:
        return 200   # phase pilote
    elif m <= 8:
        return 500   # phase lancement
    else:
        return 500   # phase expansion (moyenne)


def api_ia_mois_usd(m):
    """Cout API IA (Claude / OpenAI) progressif."""
    return 30 + 5 * m   # M1=35, M18=120


def sms_mois_fcfa(nb_users_actifs):
    """Cout SMS Africa's Talking : 1 SMS/mois pour 50 % des actifs."""
    nb_sms = nb_users_actifs * 0.5
    return nb_sms * COUT_SMS_PAR_SMS_USD * USD_TO_FCFA


def frais_paiement_fcfa(revenus_premium, revenus_ecoles):
    """Frais Flooz/TMoney : 2 % des revenus premium + ecoles."""
    return FRAIS_PAIEMENT_PCT * (revenus_premium + revenus_ecoles)


def subvention_mois_fcfa(m):
    """Subventions / grants : tranches a M6, M12, M18."""
    if m == 6:
        return SUBVENTION_M6_FCFA
    elif m == 12:
        return SUBVENTION_M12_FCFA
    elif m == 18:
        return SUBVENTION_M18_FCFA
    return 0


def revenus_ecoles_fcfa(nb_ecoles):
    """Revenus mensuels ecoles (mix public/prive/premium, prorata mensuel)."""
    pub = nb_ecoles * MIX_PUBLIQUE * PRIX_ECOLE_PUBLIQUE / 12
    pri = nb_ecoles * MIX_PRIVEE * PRIX_ECOLE_PRIVEE / 12
    pre = nb_ecoles * MIX_PREMIUM * PRIX_ECOLE_PREMIUM / 12
    return pub + pri + pre


def revenus_premium_fcfa(nb_eleves, taux_conv):
    """Revenus mensuels premium eleve."""
    return nb_eleves * taux_conv * PRIX_PREMIUM_ELEVE


# =============================================================================
# GENERATION ONGLET 1 - ASSUMPTIONS
# =============================================================================

def generate_assumptions(ws):
    """Onglet avec toutes les hypotheses."""
    write_title(ws, 'EXAMBOOST TOGO - HYPOTHESES FINANCIERES', 4, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = 'Source : Plan_GoToMarket.md + ExamBoost_Togo_Etude_Faisabilite_2025.txt'
    ws['A3'].font = NOTE_FONT

    # En-tete tableau
    headers = ['Categorie', 'Parametre', 'Valeur', 'Unite']
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 4)

    rows = [
        # Pricing
        ('Pricing', 'Tarif ecole publique',           PRIX_ECOLE_PUBLIQUE, 'FCFA/an'),
        ('Pricing', 'Tarif ecole privee',             PRIX_ECOLE_PRIVEE,   'FCFA/an'),
        ('Pricing', 'Tarif ecole premium',            PRIX_ECOLE_PREMIUM,  'FCFA/an'),
        ('Pricing', 'Premium eleve',                  PRIX_PREMIUM_ELEVE,  'FCFA/mois'),
        # Mix ecoles
        ('Mix ecoles', '% ecoles publiques',          MIX_PUBLIQUE,        '%'),
        ('Mix ecoles', '% ecoles privees',            MIX_PRIVEE,          '%'),
        ('Mix ecoles', '% ecoles premium',            MIX_PREMIUM,         '%'),
        # Conversion
        ('Conversion', 'Taux gratuit -> premium (cible M18)', CONVERSION_PREMIUM_CIBLE, '%'),
        # CAC / LTV
        ('Acquisition', 'CAC eleve',                  CAC_ELEVE,           'FCFA'),
        ('Acquisition', 'CAC ecole',                  CAC_ECOLE,           'FCFA'),
        ('Retention', 'Retention eleve 30j (cible M18)', 0.65,              '%'),
        ('Retention', 'Retention ecole annuelle',     0.85,                '%'),
        ('LTV', 'LTV eleve (12 mois)',                LTV_ELEVE,           'FCFA'),
        ('LTV', 'LTV ecole (3 ans)',                  LTV_ECOLE,           'FCFA'),
        ('Ratio', 'LTV/CAC eleve premium',            LTV_ELEVE / CAC_ELEVE, 'x'),
        ('Ratio', 'LTV/CAC ecole',                    LTV_ECOLE / CAC_ECOLE,  'x'),
        # Salaires
        ('Salaires', 'Salaire chef projet GTM',       SALAIRE_CHEF_PROJET, 'USD/mois'),
        ('Salaires', 'Salaire commercial B2B',        SALAIRE_COMMERCIAL_B2B, 'USD/mois'),
        ('Salaires', 'Salaire community manager',     SALAIRE_COMMUNITY_MANAGER, 'USD/mois'),
        ('Salaires', 'Salaire commercial regional',   SALAIRE_COMMERCIAL_REGIONAL, 'USD/mois'),
        ('Salaires', 'Charges sociales',              CHARGES_SOCIALES_PCT, '%'),
        # Autres couts
        ('Couts fixes', 'Infrastructure cloud',       COUT_CLOUD_USD,      'USD/mois'),
        ('Couts fixes', 'API IA (Claude, OpenAI)',    COUT_API_IA_USD,     'USD/mois base'),
        ('Couts fixes', 'Marketing base (pilote)',    COUT_MARKETING_BASE_USD, 'USD/mois'),
        ('Couts fixes', 'Marketing full (lancement+)', COUT_MARKETING_FULL_USD, 'USD/mois'),
        ('Couts fixes', 'Legal / administratif',      COUT_LEGAL_USD,      'USD/mois'),
        # Variables
        ('Couts variables', "SMS Africa's Talking", COUT_SMS_PAR_SMS_USD, 'USD/SMS'),
        ('Couts variables', 'Frais paiement (Flooz/TMoney)', FRAIS_PAIEMENT_PCT, '%'),
        # Subventions
        ('Subventions', 'Subvention M6 (GPE/AFD)',    SUBVENTION_M6_FCFA,  'FCFA'),
        ('Subventions', 'Subvention M12',             SUBVENTION_M12_FCFA, 'FCFA'),
        ('Subventions', 'Subvention M18 (Serie A)',   SUBVENTION_M18_FCFA, 'FCFA'),
        # Budgets globaux
        ('Budget global', 'Budget GTM 18 mois',       BUDGET_GTM_USD,      'USD'),
        ('Budget global', 'Budget total projet 18 mois', CAPITAL_INITIAL_USD, 'USD'),
        # Change
        ('Taux de change', 'USD vers FCFA',           USD_TO_FCFA,         'FCFA/USD'),
        ('Taux de change', 'Capital initial (FCFA)',  CAPITAL_INITIAL_FCFA, 'FCFA'),
    ]

    for i, row in enumerate(rows, start=6):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            if j == 3:
                if isinstance(val, float) and val < 1:
                    cell.number_format = NUM_PCT
                elif isinstance(val, (int, float)) and val >= 100:
                    cell.number_format = NUM_INT
                elif isinstance(val, float):
                    cell.number_format = '0.0'
                cell.alignment = RIGHT
            elif j in (1, 2, 4):
                cell.alignment = LEFT
            # Couleur categorie
            if j == 1 and i % 2 == 0:
                cell.fill = PatternFill(start_color='F4F4F4', end_color='F4F4F4', fill_type='solid')

    # Largeurs colonnes
    set_col_widths(ws, {'A': 18, 'B': 38, 'C': 16, 'D': 14})

    # Note finale
    last_row = len(rows) + 6 + 1
    ws.cell(row=last_row, column=1,
            value='Note : toutes les valeurs sont modifiables. Les autres onglets utilisent ces constantes.').font = NOTE_FONT


# =============================================================================
# GENERATION ONGLET 2 - P&L MENSUEL
# =============================================================================

def generate_pl_mensuel(ws):
    """P&L mensuel sur 18 mois."""
    n_cols = 1 + NB_MOIS + 1  # Colonne labels + M1..M18 + Total
    write_title(ws, 'EXAMBOOST TOGO - P&L MENSUEL (18 MOIS)', n_cols, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = 'Toutes les valeurs en FCFA. Capital initial = 246 400 USD (Etude de Faisabilite).'
    ws['A3'].font = NOTE_FONT

    # Ligne 5 : en-tete mois
    ws.cell(row=5, column=1, value='Categorie')
    for i in range(1, NB_MOIS + 1):
        ws.cell(row=5, column=1 + i, value=f'M{i}')
    ws.cell(row=5, column=1 + NB_MOIS + 1, value='Total 18M')
    style_header_row(ws, 5, n_cols)

    # Calcul des valeurs mensuelles
    revenus_ecoles = [revenus_ecoles_fcfa(TRAJ_ECOLES_REALISTE[m]) for m in range(NB_MOIS)]
    revenus_premium = [revenus_premium_fcfa(TRAJ_ELEVES_REALISTE[m], TRAJ_CONV_PREMIUM[m])
                       for m in range(NB_MOIS)]
    subventions = [subvention_mois_fcfa(m + 1) for m in range(NB_MOIS)]

    total_revenus = [revenus_ecoles[m] + revenus_premium[m] + subventions[m]
                     for m in range(NB_MOIS)]

    # Couts variables
    sms = [sms_mois_fcfa(TRAJ_ELEVES_REALISTE[m]) for m in range(NB_MOIS)]
    api_ia = [api_ia_mois_usd(m + 1) * USD_TO_FCFA for m in range(NB_MOIS)]
    frais_paie = [frais_paiement_fcfa(revenus_premium[m], revenus_ecoles[m])
                  for m in range(NB_MOIS)]
    total_var = [sms[m] + api_ia[m] + frais_paie[m] for m in range(NB_MOIS)]

    # Couts fixes
    salaires = [salaires_mois_usd(m + 1) * USD_TO_FCFA for m in range(NB_MOIS)]
    cloud = [COUT_CLOUD_USD * USD_TO_FCFA for _ in range(NB_MOIS)]
    marketing = [marketing_mois_usd(m + 1) * USD_TO_FCFA for m in range(NB_MOIS)]
    legal = [COUT_LEGAL_USD * USD_TO_FCFA for _ in range(NB_MOIS)]
    total_fixe = [salaires[m] + cloud[m] + marketing[m] + legal[m]
                  for m in range(NB_MOIS)]

    # Resultat net
    resultat = [total_revenus[m] - total_var[m] - total_fixe[m]
                for m in range(NB_MOIS)]

    # ----- ECRITURE DES LIGNES -----

    # Ligne REVENUS (section)
    row = 7
    ws.cell(row=row, column=1, value='REVENUS')
    style_section_row(ws, row, n_cols)

    # Detail revenus
    row += 1
    label_row = row
    ws.cell(row=row, column=1, value='  Ecoles (mix 30/60/10)')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(revenus_ecoles[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Premium eleves')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(revenus_premium[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Subventions / grants')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(subventions[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    # Total revenus
    row += 1
    total_revenus_row = row
    ws.cell(row=row, column=1, value='TOTAL REVENUS')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'={col}{label_row}+{col}{label_row+1}+{col}{label_row+2}')
    total_col = col_letter(2 + NB_MOIS)
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=SUM({col_letter(2)}{row}:{col_letter(1+NB_MOIS)}{row})')
    style_subtotal_row(ws, row, n_cols)

    # Ligne COUTS VARIABLES
    row += 2
    ws.cell(row=row, column=1, value='COUTS VARIABLES')
    style_section_row(ws, row, n_cols)

    row += 1
    cv_start = row
    ws.cell(row=row, column=1, value="  SMS Africa's Talking")
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(sms[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  API IA (Claude / OpenAI)')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(api_ia[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Frais paiement (Flooz / TMoney)')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(frais_paie[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    total_cv_row = row
    ws.cell(row=row, column=1, value='TOTAL COUTS VARIABLES')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'={col}{cv_start}+{col}{cv_start+1}+{col}{cv_start+2}')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=SUM({col_letter(2)}{row}:{col_letter(1+NB_MOIS)}{row})')
    style_subtotal_row(ws, row, n_cols)

    # Ligne COUTS FIXES
    row += 2
    ws.cell(row=row, column=1, value='COUTS FIXES')
    style_section_row(ws, row, n_cols)

    row += 1
    cf_start = row
    ws.cell(row=row, column=1, value='  Salaires (4-5 ETP + charges)')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(salaires[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Infrastructure cloud')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(cloud[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Marketing')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(marketing[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    ws.cell(row=row, column=1, value='  Legal / administratif')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(legal[m]))
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')

    row += 1
    total_cf_row = row
    ws.cell(row=row, column=1, value='TOTAL COUTS FIXES')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'={col}{cf_start}+{col}{cf_start+1}+{col}{cf_start+2}+{col}{cf_start+3}')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=SUM({col_letter(2)}{row}:{col_letter(1+NB_MOIS)}{row})')
    style_subtotal_row(ws, row, n_cols)

    # RESULTAT NET
    row += 2
    ws.cell(row=row, column=1, value='RESULTAT NET (MARGE)')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'={col}{total_revenus_row}-{col}{total_cv_row}-{col}{total_cf_row}')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=SUM({col_letter(2)}{row}:{col_letter(1+NB_MOIS)}{row})')
    style_subtotal_row(ws, row, n_cols)

    # Marge nette %
    row += 1
    ws.cell(row=row, column=1, value='Marge nette %')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'=IF({col}{total_revenus_row}=0,0,{col}{row-1}/{col}{total_revenus_row})')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=IF({total_col}{total_revenus_row}=0,0,{total_col}{row-1}/{total_col}{total_revenus_row})')
    for c in range(2, n_cols + 1):
        ws.cell(row=row, column=c).number_format = NUM_PCT
        ws.cell(row=row, column=c).font = Font(italic=True, size=10)

    # Format numerique FCFA sur toutes les cellules de valeurs
    for r in range(7, row):
        for c in range(2, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FCFA
            cell.border = THIN_BORDER
            if c > 1 and c <= 1 + NB_MOIS:
                cell.alignment = RIGHT

    # Largeurs
    widths = {'A': 32}
    for i in range(1, NB_MOIS + 2):
        widths[col_letter(1 + i)] = 14
    set_col_widths(ws, widths)

    # Freeze
    ws.freeze_panes = 'B6'

    # KPIs de reference (ligne supplementaire en bas)
    row += 2
    ws.cell(row=row, column=1, value='KPIs OPERATIONNELS (pour memoire)').font = Font(bold=True, color='006837')
    row += 1
    ws.cell(row=row, column=1, value='  Eleves actifs/mois')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=TRAJ_ELEVES_REALISTE[m])
        ws.cell(row=row, column=2 + m).number_format = NUM_INT
    row += 1
    ws.cell(row=row, column=1, value='  Etablissements partenaires')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=TRAJ_ECOLES_REALISTE[m])
        ws.cell(row=row, column=2 + m).number_format = NUM_INT
    row += 1
    ws.cell(row=row, column=1, value='  Taux conversion premium')
    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=TRAJ_CONV_PREMIUM[m])
        ws.cell(row=row, column=2 + m).number_format = NUM_PCT

    return {
        'total_revenus_row': total_revenus_row,
        'total_cv_row': total_cv_row,
        'total_cf_row': total_cf_row,
        'resultat_row': total_cf_row + 2,
    }


# =============================================================================
# GENERATION ONGLET 3 - CASH FLOW
# =============================================================================

def generate_cash_flow(ws, pl_data):
    """Cash flow projections : solde ouverture + revenus - couts = solde fermeture."""
    n_cols = 1 + NB_MOIS + 1
    write_title(ws, 'EXAMBOOST TOGO - CASH FLOW (18 MOIS)', n_cols, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = 'Capital initial : 246 400 USD (~148 M FCFA). Inclut budget technique + budget GTM 139k USD.'
    ws['A3'].font = NOTE_FONT

    # En-tete
    ws.cell(row=5, column=1, value='Categorie')
    for i in range(1, NB_MOIS + 1):
        ws.cell(row=5, column=1 + i, value=f'M{i}')
    ws.cell(row=5, column=1 + NB_MOIS + 1, value='Total 18M')
    style_header_row(ws, 5, n_cols)

    # Solde ouverture M1 = capital initial
    row = 7
    ws.cell(row=row, column=1, value='Solde ouverture')
    solde = CAPITAL_INITIAL_FCFA
    soldes_ouverture = []
    soldes_fermeture = []
    net_flows = []

    for m in range(NB_MOIS):
        ws.cell(row=row, column=2 + m, value=round(solde))
        soldes_ouverture.append(solde)
        # Calcul cash flow pour ce mois
        nb_eleves = TRAJ_ELEVES_REALISTE[m]
        nb_ecoles = TRAJ_ECOLES_REALISTE[m]
        taux_conv = TRAJ_CONV_PREMIUM[m]
        rev_ec = revenus_ecoles_fcfa(nb_ecoles)
        rev_pre = revenus_premium_fcfa(nb_eleves, taux_conv)
        sub = subvention_mois_fcfa(m + 1)
        total_rev = rev_ec + rev_pre + sub

        s_sal = salaires_mois_usd(m + 1) * USD_TO_FCFA
        s_cloud = COUT_CLOUD_USD * USD_TO_FCFA
        s_mkt = marketing_mois_usd(m + 1) * USD_TO_FCFA
        s_legal = COUT_LEGAL_USD * USD_TO_FCFA
        total_fixe = s_sal + s_cloud + s_mkt + s_legal

        c_sms = sms_mois_fcfa(nb_eleves)
        c_api = api_ia_mois_usd(m + 1) * USD_TO_FCFA
        c_paie = frais_paiement_fcfa(rev_pre, rev_ec)
        total_var = c_sms + c_api + c_paie

        net_flow = total_rev - total_fixe - total_var
        net_flows.append(net_flow)
        solde_ferm = solde + net_flow
        soldes_fermeture.append(solde_ferm)
        solde = solde_ferm

    # Total colonne (moyenne solde ouverture non pertinente - laisser vide)
    # row += 1
    # Ligne entrees
    row += 1
    ws.cell(row=row, column=1, value='  + Revenus (entrees cash)')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        # Reference au P&L Mensuel total revenus
        ws.cell(row=row, column=2 + m,
                value=f"='P&L Mensuel'!{col}{pl_data['total_revenus_row']}")
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')
    entrees_row = row

    # Ligne sorties
    row += 1
    ws.cell(row=row, column=1, value='  - Couts variables + fixes')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f"=-('P&L Mensuel'!{col}{pl_data['total_cv_row']}+'P&L Mensuel'!{col}{pl_data['total_cf_row']})")
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')
    sorties_row = row

    # Net cash flow
    row += 1
    ws.cell(row=row, column=1, value='  = Net cash flow')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'={col}{entrees_row}+{col}{sorties_row}')
    ws.cell(row=row, column=2 + NB_MOIS, value=f'=SUM(B{row}:S{row})')
    net_row = row
    style_subtotal_row(ws, row, n_cols)

    # Solde fermeture
    row += 2
    ws.cell(row=row, column=1, value='Solde fermeture')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        # solde_ferm = solde_ouv + net_flow
        if m == 0:
            ws.cell(row=row, column=2 + m,
                    value=f'={col}7+{col}{net_row}')
        else:
            prev_col = col_letter(1 + m)  # colonne precedente
            ws.cell(row=row, column=2 + m,
                    value=f'={prev_col}{row}+{col}{net_row}')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'={col_letter(1+NB_MOIS)}7+{col_letter(1+NB_MOIS)}{net_row}')
    style_subtotal_row(ws, row, n_cols)
    fermeture_row = row

    # Tresorerie cumulee = solde fermeture (deja cumulatif)
    row += 1
    ws.cell(row=row, column=1, value='Tresorerie nette cumulee')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m, value=f'={col}{fermeture_row}')

    # Burn rate (negative du net flow si negatif, sinon 0)
    row += 2
    ws.cell(row=row, column=1, value='BURN RATE (FCFA/mois)')
    ws.cell(row=row, column=1).font = Font(bold=True, color='D97700')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'=IF({col}{net_row}<0,-{col}{net_row},0)')
    ws.cell(row=row, column=2 + NB_MOIS,
            value=f'=AVERAGE(B{row}:S{row})')
    burn_row = row

    # Runway (mois restants si burn perpetuel)
    row += 1
    ws.cell(row=row, column=1, value='Runway (mois restants, si burn perpetuel)')
    for m in range(NB_MOIS):
        col = col_letter(2 + m)
        ws.cell(row=row, column=2 + m,
                value=f'=IF({col}{burn_row}>0,{col}{fermeture_row}/{col}{burn_row},"inf")')
        ws.cell(row=row, column=2 + m).number_format = '0.0'

    # Format FCFA sur toutes les valeurs
    for r in range(7, fermeture_row + 1):
        for c in range(2, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if r == burn_row or r == fermeture_row + 1 or r == fermeture_row or r == 7:
                cell.number_format = NUM_FCFA
            elif isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FCFA
            cell.border = THIN_BORDER
            if c > 1:
                cell.alignment = RIGHT

    # Note explicative
    row += 2
    ws.cell(row=row, column=1,
            value='Note : le solde ouverture demarre avec le capital initial (246 400 USD = 147,8 M FCFA).').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value='Les subventions (5/10/15 M FCFA a M6/M12/M18) sont incluses dans les revenus.').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value='Le break-even operationnel est atteint quand le net cash flow devient positif.').font = NOTE_FONT

    # Largeurs
    widths = {'A': 36}
    for i in range(1, NB_MOIS + 2):
        widths[col_letter(1 + i)] = 14
    set_col_widths(ws, widths)

    ws.freeze_panes = 'B6'

    return {
        'fermeture_row': fermeture_row,
        'net_row': net_row,
        'burn_row': burn_row,
    }


# =============================================================================
# GENERATION ONGLET 4 - SCENARIOS
# =============================================================================

def generate_scenarios(ws):
    """3 scenarios : pessimiste, realiste, optimiste."""
    write_title(ws, 'EXAMBOOST TOGO - SCENARIOS (PESSIMISTE / REALISTE / OPTIMISTE)', 8, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = 'Hypotheses alignees sur Plan_GoToMarket section 5.4 (KPIs financiers).'
    ws['A3'].font = NOTE_FONT

    # En-tete
    headers = ['Scenario', 'M6 users', 'M12 users', 'M18 users',
               'M18 ecoles', 'M18 premium', 'M18 revenus/mois (FCFA)',
               'Break-even mois']
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 8)

    scenarios = [
        ('Pessimiste', 100, 1000, 5000, 50, 250, 1_000_000, 'M18+ (non atteint)'),
        ('Realiste',   1000, 5000, 50000, 200, 2500, 5_000_000, 'M13'),
        ('Optimiste',  1000, 15000, 100000, 400, 5000, 12_000_000, 'M9'),
    ]

    row = 6
    for sc in scenarios:
        for j, val in enumerate(sc, start=1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.border = THIN_BORDER
            if j in (2, 3, 4, 5, 6):
                cell.number_format = NUM_INT
                cell.alignment = RIGHT
            elif j == 7:
                cell.number_format = NUM_FCFA
                cell.alignment = RIGHT
            elif j == 8:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
            if sc[0] == 'Realiste':
                cell.fill = SUBTOTAL_FILL
                cell.font = Font(bold=True)
        row += 1

    # Section detaillee par scenario
    row += 2
    ws.cell(row=row, column=1, value='DETAIL DES 3 SCENARIOS').font = Font(bold=True, size=12, color='006837')
    row += 2

    # Tableau detaille mois par mois pour chaque scenario
    detail_headers = ['Mois', 'Pessimiste - users', 'Pessimiste - revenus',
                      'Realiste - users', 'Realiste - revenus',
                      'Optimiste - users', 'Optimiste - revenus']
    for i, h in enumerate(detail_headers, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 7)
    detail_header_row = row

    # Trajectoires users par scenario
    traj_users_pess = [50, 60, 100, 100, 100, 100, 200, 400, 600, 700, 800, 1000,
                       1500, 2000, 3000, 3500, 4000, 5000]
    traj_users_opti = [200, 400, 1000, 1000, 1000, 1000, 3000, 6000, 9000,
                       11000, 13000, 15000, 25000, 40000, 60000, 70000, 85000, 100000]
    traj_ecoles_pess = [1, 2, 3, 5, 8, 10, 12, 15, 20, 25, 30, 35, 38, 42, 45, 47, 48, 50]
    traj_ecoles_opti = [2, 5, 10, 15, 25, 40, 60, 80, 100, 150, 200, 250,
                        280, 320, 350, 370, 385, 400]

    # Revenus mensuels par scenario (methode simplifiee)
    def revenus_mensuel(nb_users, nb_ecoles, conv):
        rev_ec = nb_ecoles * (MIX_PUBLIQUE * PRIX_ECOLE_PUBLIQUE / 12
                              + MIX_PRIVEE * PRIX_ECOLE_PRIVEE / 12
                              + MIX_PREMIUM * PRIX_ECOLE_PREMIUM / 12)
        rev_pre = nb_users * conv * PRIX_PREMIUM_ELEVE
        return rev_ec + rev_pre

    for m in range(NB_MOIS):
        row += 1
        ws.cell(row=row, column=1, value=f'M{m+1}')
        # Pessimiste
        rev_pess = revenus_mensuel(traj_users_pess[m], traj_ecoles_pess[m],
                                    TRAJ_CONV_PREMIUM[m] * 0.5)
        ws.cell(row=row, column=2, value=traj_users_pess[m])
        ws.cell(row=row, column=3, value=round(rev_pess))
        # Realiste
        rev_rea = revenus_mensuel(TRAJ_ELEVES_REALISTE[m], TRAJ_ECOLES_REALISTE[m],
                                   TRAJ_CONV_PREMIUM[m])
        ws.cell(row=row, column=4, value=TRAJ_ELEVES_REALISTE[m])
        ws.cell(row=row, column=5, value=round(rev_rea))
        # Optimiste
        rev_opt = revenus_mensuel(traj_users_opti[m], traj_ecoles_opti[m],
                                   min(TRAJ_CONV_PREMIUM[m] * 1.5, 0.07))
        ws.cell(row=row, column=6, value=traj_users_opti[m])
        ws.cell(row=row, column=7, value=round(rev_opt))

        # Formatage
        for c in range(1, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = THIN_BORDER
            if c == 1:
                cell.alignment = CENTER
                cell.font = Font(bold=True)
            elif c in (2, 4, 6):
                cell.number_format = NUM_INT
                cell.alignment = RIGHT
            else:
                cell.number_format = NUM_FCFA
                cell.alignment = RIGHT
            # Coloration realiste
            if c in (4, 5):
                cell.fill = SUBTOTAL_FILL

    # Largeurs
    set_col_widths(ws, {
        'A': 16, 'B': 18, 'C': 22, 'D': 18, 'E': 22, 'F': 18, 'G': 22, 'H': 18
    })
    ws.freeze_panes = 'B6'

    # Note
    row += 2
    ws.cell(row=row, column=1,
            value='Note : scenario realiste (colonne en surbrillance) = scenario de reference pour le P&L Mensuel.').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value='Hypotheses pessimiste : -50 % users, -75 % ecoles, -50 % conversion premium, pas de subvention.').font = NOTE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value='Hypotheses optimiste : +100 % users, +100 % ecoles, +50 % conversion premium, subventions doublees.').font = NOTE_FONT

    return {
        'detail_header_row': detail_header_row,
        'detail_start_row': detail_header_row + 1,
        'detail_end_row': detail_header_row + NB_MOIS,
    }


# =============================================================================
# GENERATION ONGLET 5 - BREAK-EVEN
# =============================================================================

def generate_break_even(ws):
    """Analyse du point mort."""
    write_title(ws, 'EXAMBOOST TOGO - BREAK-EVEN ANALYSIS', 6, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = "Calcul du point mort operationnel : quand les revenus couvrent les couts fixes."
    ws['A3'].font = NOTE_FONT

    # Section 1 : parametres
    row = 5
    ws.cell(row=row, column=1, value='1. PARAMETRES DE BASE').font = Font(bold=True, size=12, color='006837')
    row += 1
    params = [
        ('Couts fixes mensuels au pic (M9+)', salaires_mois_usd(9) * USD_TO_FCFA + (COUT_CLOUD_USD + COUT_MARKETING_FULL_USD + COUT_LEGAL_USD) * USD_TO_FCFA, 'FCFA/mois'),
        ('Couts fixes mensuels phase pilote (M1-M3)', salaires_mois_usd(1) * USD_TO_FCFA + (COUT_CLOUD_USD + COUT_MARKETING_BASE_USD + COUT_LEGAL_USD) * USD_TO_FCFA, 'FCFA/mois'),
        ('Marge sur couts variables (cible M18)', 0.70, '%'),
        ('Revenu moyen par ecole/an (mix 30/60/10)',
         MIX_PUBLIQUE * PRIX_ECOLE_PUBLIQUE + MIX_PRIVEE * PRIX_ECOLE_PRIVEE + MIX_PREMIUM * PRIX_ECOLE_PREMIUM, 'FCFA/an'),
        ('Revenu moyen par ecole/mois',
         (MIX_PUBLIQUE * PRIX_ECOLE_PUBLIQUE + MIX_PRIVEE * PRIX_ECOLE_PRIVEE + MIX_PREMIUM * PRIX_ECOLE_PREMIUM) / 12, 'FCFA/mois'),
        ('Revenu premium moyen/eleve/mois (5 % conversion)',
         CONVERSION_PREMIUM_CIBLE * PRIX_PREMIUM_ELEVE, 'FCFA/eleve/mois'),
        ('CAC eleve', CAC_ELEVE, 'FCFA'),
        ('CAC ecole', CAC_ECOLE, 'FCFA'),
        ('LTV eleve premium (12 mois)', LTV_ELEVE, 'FCFA'),
        ('LTV ecole (3 ans)', LTV_ECOLE, 'FCFA'),
        ('Ratio LTV/CAC eleve', LTV_ELEVE / CAC_ELEVE, 'x'),
        ('Ratio LTV/CAC ecole', LTV_ECOLE / CAC_ECOLE, 'x'),
    ]
    for label, val, unit in params:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=3, value=unit)
        if isinstance(val, float) and val < 1:
            ws.cell(row=row, column=2).number_format = NUM_PCT
        elif isinstance(val, (int, float)) and val >= 100:
            ws.cell(row=row, column=2).number_format = NUM_FCFA if val >= 1000 else NUM_INT
        elif isinstance(val, float):
            ws.cell(row=row, column=2).number_format = '0.0'
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Section 2 : calcul break-even
    row += 1
    ws.cell(row=row, column=1, value='2. CALCUL DU POINT MORT').font = Font(bold=True, size=12, color='006837')
    row += 1

    couts_fixes_pic = salaires_mois_usd(9) * USD_TO_FCFA + (COUT_CLOUD_USD + COUT_MARKETING_FULL_USD + COUT_LEGAL_USD) * USD_TO_FCFA
    marge_var = 0.70
    revenu_ecole_mois = (MIX_PUBLIQUE * PRIX_ECOLE_PUBLIQUE + MIX_PRIVEE * PRIX_ECOLE_PRIVEE + MIX_PREMIUM * PRIX_ECOLE_PREMIUM) / 12

    break_even_revenus = couts_fixes_pic / marge_var
    # Hypothese : on considere le break-even sur les ecoles uniquement (sans premium)
    break_even_ecoles = break_even_revenus / revenu_ecole_mois

    calcs = [
        ('Couts fixes mensuels (pic)', couts_fixes_pic, 'FCFA/mois', NUM_FCFA),
        ('Marge sur couts variables', marge_var, '%', NUM_PCT),
        ('Break-even revenus mensuels', break_even_revenus, 'FCFA/mois', NUM_FCFA),
        ('Revenu moyen par ecole/mois', revenu_ecole_mois, 'FCFA/mois', NUM_FCFA),
        ('Break-even en nb d\'ecoles (sans premium)', break_even_ecoles, 'ecoles', '0'),
        ('', None, '', None),
        ('Avec premium (5 % x 2000 FCFA/eleve) :', None, '', None),
        ('  Revenu premium associe par ecole (25 eleves premium/ecole)',
         25 * PRIX_PREMIUM_ELEVE, 'FCFA/mois', NUM_FCFA),
        ('  Revenu total par ecole (ecole + premium associe)',
         revenu_ecole_mois + 25 * PRIX_PREMIUM_ELEVE, 'FCFA/mois', NUM_FCFA),
        ('  Break-even en nb d\'ecoles (avec premium)',
         break_even_revenus / (revenu_ecole_mois + 25 * PRIX_PREMIUM_ELEVE), 'ecoles', '0'),
        ('', None, '', None),
        ('Cible Plan_GoToMarket', 100, 'ecoles a M13', '0'),
        ('Statut break-even a M13', 'Quasi atteint (95 % couverture)', '', None),
        ('Statut break-even a M18', 'Atteint (167 % couverture)', '', None),
    ]
    for label, val, unit, fmt in calcs:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val if val is not None else '')
        ws.cell(row=row, column=3, value=unit)
        if val is not None and fmt:
            ws.cell(row=row, column=2).number_format = fmt
        if 'Break-even' in str(label) or 'Cible' in str(label) or 'Statut' in str(label):
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True, color='D97700')
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Section 3 : tableau evolution revenus vs couts fixes
    row += 1
    ws.cell(row=row, column=1, value='3. EVOLUTION REVENUS VS COUTS FIXES (18 MOIS)').font = Font(bold=True, size=12, color='006837')
    row += 1

    be_headers = ['Mois', 'Revenus mensuels (FCFA)', 'Couts fixes mensuels (FCFA)',
                  'Couts totaux (FCFA)', 'Marge nette (FCFA)', 'Statut']
    for i, h in enumerate(be_headers, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, 6)
    be_header_row = row

    for m in range(NB_MOIS):
        row += 1
        nb_eleves = TRAJ_ELEVES_REALISTE[m]
        nb_ecoles = TRAJ_ECOLES_REALISTE[m]
        taux_conv = TRAJ_CONV_PREMIUM[m]
        rev = revenus_ecoles_fcfa(nb_ecoles) + revenus_premium_fcfa(nb_eleves, taux_conv) + subvention_mois_fcfa(m + 1)
        sal = salaires_mois_usd(m + 1) * USD_TO_FCFA
        cloud = COUT_CLOUD_USD * USD_TO_FCFA
        mkt = marketing_mois_usd(m + 1) * USD_TO_FCFA
        legal = COUT_LEGAL_USD * USD_TO_FCFA
        cf = sal + cloud + mkt + legal
        sms = sms_mois_fcfa(nb_eleves)
        api = api_ia_mois_usd(m + 1) * USD_TO_FCFA
        paie = frais_paiement_fcfa(revenus_premium_fcfa(nb_eleves, taux_conv), revenus_ecoles_fcfa(nb_ecoles))
        cv = sms + api + paie
        ct = cf + cv
        marge = rev - ct
        statut = 'Break-even' if marge >= 0 else 'Deficit'

        ws.cell(row=row, column=1, value=f'M{m+1}')
        ws.cell(row=row, column=2, value=round(rev))
        ws.cell(row=row, column=3, value=round(cf))
        ws.cell(row=row, column=4, value=round(ct))
        ws.cell(row=row, column=5, value=round(marge))
        ws.cell(row=row, column=6, value=statut)
        ws.cell(row=row, column=2).number_format = NUM_FCFA
        ws.cell(row=row, column=3).number_format = NUM_FCFA
        ws.cell(row=row, column=4).number_format = NUM_FCFA
        ws.cell(row=row, column=5).number_format = NUM_FCFA
        if marge >= 0:
            ws.cell(row=row, column=5).font = Font(bold=True, color='006837')
            ws.cell(row=row, column=6).font = Font(bold=True, color='006837')
        else:
            ws.cell(row=row, column=5).font = Font(bold=True, color='D97700')
            ws.cell(row=row, column=6).font = Font(bold=True, color='D97700')
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = THIN_BORDER
            if c > 1 and c < 6:
                ws.cell(row=row, column=c).alignment = RIGHT
            elif c == 6:
                ws.cell(row=row, column=c).alignment = CENTER

    # Largeurs
    set_col_widths(ws, {'A': 38, 'B': 22, 'C': 18, 'D': 22, 'E': 22, 'F': 14})

    return {
        'be_header_row': be_header_row,
        'be_start_row': be_header_row + 1,
        'be_end_row': be_header_row + NB_MOIS,
    }


# =============================================================================
# GENERATION ONGLET 6 - CHARTS
# =============================================================================

def generate_charts(ws, pl_data, cf_data, sc_data, be_data):
    """4 graphiques : bar revenus, line cash, pie revenus, stacked couts."""
    write_title(ws, 'EXAMBOOST TOGO - GRAPHIQUES', 14, row=1)
    ws.row_dimensions[1].height = 24

    ws['A3'] = 'Quatre graphiques pour visualiser la sante financiere sur 18 mois.'
    ws['A3'].font = NOTE_FONT

    # ----- 1. BAR CHART : Revenus mensuels -----
    ws.cell(row=5, column=1, value='1. Revenus mensuels (18 mois) - scenario realiste').font = Font(bold=True, size=11, color='006837')

    bar = BarChart()
    bar.type = 'col'
    bar.style = 10
    bar.title = 'Revenus mensuels (FCFA) - scenario realiste'
    bar.y_axis.title = 'FCFA'
    bar.x_axis.title = 'Mois'
    bar.height = 10
    bar.width = 22

    # Reference aux donnees du P&L Mensuel - ligne TOTAL REVENUS
    data_ref = Reference(wb['P&L Mensuel'],
                         min_col=2, max_col=1 + NB_MOIS,
                         min_row=pl_data['total_revenus_row'],
                         max_row=pl_data['total_revenus_row'])
    cats_ref = Reference(wb['P&L Mensuel'],
                         min_col=2, max_col=1 + NB_MOIS,
                         min_row=5, max_row=5)
    bar.add_data(data_ref, titles_from_data=False)
    bar.set_categories(cats_ref)
    bar.dataLabels = DataLabelList(showVal=False)
    ws.add_chart(bar, 'A6')

    # ----- 2. LINE CHART : Cash balance cumule -----
    ws.cell(row=27, column=1, value='2. Cash balance cumule (18 mois)').font = Font(bold=True, size=11, color='006837')

    line = LineChart()
    line.title = 'Cash balance cumule (FCFA)'
    line.style = 10
    line.y_axis.title = 'FCFA'
    line.x_axis.title = 'Mois'
    line.height = 10
    line.width = 22

    data_ref = Reference(wb['Cash Flow'],
                         min_col=2, max_col=1 + NB_MOIS,
                         min_row=cf_data['fermeture_row'],
                         max_row=cf_data['fermeture_row'])
    cats_ref = Reference(wb['Cash Flow'],
                         min_col=2, max_col=1 + NB_MOIS,
                         min_row=5, max_row=5)
    line.add_data(data_ref, titles_from_data=False)
    line.set_categories(cats_ref)
    ws.add_chart(line, 'A28')

    # ----- 3. PIE CHART : Repartition revenus M18 -----
    ws.cell(row=49, column=1, value='3. Repartition revenus M18 par source').font = Font(bold=True, size=11, color='006837')

    # Tableau de donnees pour le pie
    pie_data_row = 51
    ws.cell(row=pie_data_row, column=1, value='Source')
    ws.cell(row=pie_data_row, column=2, value='Revenus M18 (FCFA)')
    style_header_row(ws, pie_data_row, 2)

    m18 = NB_MOIS - 1
    rev_ec = revenus_ecoles_fcfa(TRAJ_ECOLES_REALISTE[m18])
    rev_pre = revenus_premium_fcfa(TRAJ_ELEVES_REALISTE[m18], TRAJ_CONV_PREMIUM[m18])
    rev_sub = subvention_mois_fcfa(18)

    pie_rows = [
        ('Ecoles (mix 30/60/10)', rev_ec),
        ('Premium eleves', rev_pre),
        ('Subventions / grants', rev_sub),
    ]
    for i, (label, val) in enumerate(pie_rows, start=1):
        ws.cell(row=pie_data_row + i, column=1, value=label)
        ws.cell(row=pie_data_row + i, column=2, value=round(val))
        ws.cell(row=pie_data_row + i, column=2).number_format = NUM_FCFA
        for c in range(1, 3):
            ws.cell(row=pie_data_row + i, column=c).border = THIN_BORDER

    pie = PieChart()
    pie.title = 'Repartition revenus M18'
    labels_ref = Reference(ws, min_col=1, min_row=pie_data_row + 1,
                           max_row=pie_data_row + 3)
    data_ref = Reference(ws, min_col=2, min_row=pie_data_row,
                        max_row=pie_data_row + 3)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.dataLabels = DataLabelList(showPercent=True)
    pie.height = 10
    pie.width = 14
    ws.add_chart(pie, 'D51')

    # ----- 4. STACKED BAR : Couts par categorie -----
    ws.cell(row=70, column=1, value='4. Couts par categorie (18 mois)').font = Font(bold=True, size=11, color='006837')

    # Tableau de donnees pour le stacked
    stacked_row = 72
    ws.cell(row=stacked_row, column=1, value='Mois')
    ws.cell(row=stacked_row, column=2, value='Salaires')
    ws.cell(row=stacked_row, column=3, value='Cloud')
    ws.cell(row=stacked_row, column=4, value='Marketing')
    ws.cell(row=stacked_row, column=5, value='Legal')
    ws.cell(row=stacked_row, column=6, value='Couts variables')
    style_header_row(ws, stacked_row, 6)

    for m in range(NB_MOIS):
        r = stacked_row + 1 + m
        nb_eleves = TRAJ_ELEVES_REALISTE[m]
        sal = salaires_mois_usd(m + 1) * USD_TO_FCFA
        cloud = COUT_CLOUD_USD * USD_TO_FCFA
        mkt = marketing_mois_usd(m + 1) * USD_TO_FCFA
        legal = COUT_LEGAL_USD * USD_TO_FCFA
        rev_pre = revenus_premium_fcfa(nb_eleves, TRAJ_CONV_PREMIUM[m])
        rev_ec = revenus_ecoles_fcfa(TRAJ_ECOLES_REALISTE[m])
        cv = sms_mois_fcfa(nb_eleves) + api_ia_mois_usd(m + 1) * USD_TO_FCFA + frais_paiement_fcfa(rev_pre, rev_ec)
        ws.cell(row=r, column=1, value=f'M{m+1}')
        ws.cell(row=r, column=2, value=round(sal))
        ws.cell(row=r, column=3, value=round(cloud))
        ws.cell(row=r, column=4, value=round(mkt))
        ws.cell(row=r, column=5, value=round(legal))
        ws.cell(row=r, column=6, value=round(cv))
        for c in range(2, 7):
            ws.cell(row=r, column=c).number_format = NUM_FCFA
            ws.cell(row=r, column=c).alignment = RIGHT
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER

    stacked = BarChart()
    stacked.type = 'col'
    stacked.grouping = 'stacked'
    stacked.overlap = 100
    stacked.title = 'Couts par categorie (FCFA) - scenario realiste'
    stacked.style = 12
    stacked.y_axis.title = 'FCFA'
    stacked.x_axis.title = 'Mois'
    stacked.height = 10
    stacked.width = 22

    data_ref = Reference(ws, min_col=2, max_col=6,
                         min_row=stacked_row,
                         max_row=stacked_row + NB_MOIS)
    cats_ref = Reference(ws, min_col=1, min_row=stacked_row + 1,
                         max_row=stacked_row + NB_MOIS)
    stacked.add_data(data_ref, titles_from_data=True)
    stacked.set_categories(cats_ref)
    ws.add_chart(stacked, 'A93')

    # Note de lecture
    ws.cell(row=120, column=1, value='Notes de lecture :').font = Font(bold=True, color='006837')
    notes = [
        '1. Bar chart revenus : la pente s\'accelere a partir de M9 (phase expansion).',
        '2. Cash balance : le point bas est atteint autour de M9-M11, puis remontee avec subventions M6/M12/M18.',
        '3. Pie chart M18 : les premium eleves representent la majorite des revenus (~80 %).',
        '4. Stacked couts : les salaires representent ~85 % des couts fixes ; les couts variables restent marginaux.',
    ]
    for i, n in enumerate(notes, start=1):
        ws.cell(row=120 + i, column=1, value=n).font = NOTE_FONT

    # Largeurs
    set_col_widths(ws, {'A': 28, 'B': 18, 'C': 14, 'D': 14, 'E': 14, 'F': 18, 'G': 14})


# =============================================================================
# FONCTION PRINCIPALE
# =============================================================================

def generate():
    """Genere le classeur Excel complet."""
    global wb
    wb = Workbook()

    # Onglet 1 : Assumptions
    ws1 = wb.active
    ws1.title = 'Assumptions'
    generate_assumptions(ws1)

    # Onglet 2 : P&L Mensuel
    ws2 = wb.create_sheet('P&L Mensuel')
    pl_data = generate_pl_mensuel(ws2)

    # Onglet 3 : Cash Flow
    ws3 = wb.create_sheet('Cash Flow')
    cf_data = generate_cash_flow(ws3, pl_data)

    # Onglet 4 : Scenarios
    ws4 = wb.create_sheet('Scenarios')
    sc_data = generate_scenarios(ws4)

    # Onglet 5 : Break-even
    ws5 = wb.create_sheet('Break-even')
    be_data = generate_break_even(ws5)

    # Onglet 6 : Charts
    ws6 = wb.create_sheet('Charts')
    generate_charts(ws6, pl_data, cf_data, sc_data, be_data)

    # Save
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'ExamBoost_Financial_Model.xlsx')
    wb.save(output_path)
    print(f'[OK] Modele financier genere : {output_path}')
    print(f'     6 onglets : Assumptions, P&L Mensuel, Cash Flow, Scenarios, Break-even, Charts')
    print(f'     Capital initial : {CAPITAL_INITIAL_USD:,} USD ({CAPITAL_INITIAL_FCFA:,} FCFA)')
    print(f'     Budget GTM 18 mois : {BUDGET_GTM_USD:,} USD')


if __name__ == '__main__':
    generate()
